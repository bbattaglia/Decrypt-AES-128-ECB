"""
AES (Advanced Encryption Standard) è un algoritmo di cifratura a blocchi che può utilizzare chiavi di 128, 192 o 256 bit. AES-128 utilizza chiavi di 128 bit, il che significa che la chiave utilizzata per la cifratura e la decifratura ha una lunghezza di 128 bit (16 byte).

AES-128 utilizza una tecnica di cifratura a blocchi, dove i dati in chiaro sono suddivisi in blocchi di 128 bit e ogni blocco viene cifrato indipendentemente dagli altri. In modalità ECB (Electronic CodeBook), ogni blocco di 128 bit viene cifrato indipendentemente dagli altri, utilizzando la stessa chiave di cifratura.

Il padding PKCS5Padding viene utilizzato per aggiungere un numero variabile di byte al blocco di dati in chiaro in modo che il numero totale di byte del blocco diventi un multiplo di 128 bit.

Quindi, nel caso specifico della stringa "jeVTGDnzL0P4n63NIWuEaw==" decriptata con la chiave "958BAE842AA8D07765F7EF9E5D0CDEC6", i passaggi sono i seguenti:

1. La chiave esadecimale viene convertita in una sequenza di byte.
2. La stringa "jeVTGDnzL0P4n63NIWuEaw==" viene decodificata dalla codifica Base64 in una sequenza di byte.
3. Viene inizializzato un cifrario AES-128 in modalità ECB con la chiave convertita e il padding PKCS5Padding.
4. La sequenza di byte cifrati viene decriptata utilizzando il cifrario AES-128 in modalità ECB.
5. Il testo in chiaro risultante dalla decriptazione viene convertito in una stringa utilizzando il set di caratteri UTF-8.
"""

import base64
from Crypto.Cipher import AES
import unicodedata
import csv
import xlsxwriter
import os 
import openpyxl
import datetime


KEY = "958BAE842AA8D07765F7EF9E5D0CDEC6"
columdToDecrypt = ['D','I','M'] 
inputFile = "C:\Git\Golden Goose\Decrypt AES-128-ECB\memberBindQuery.csv"
outputFile = "C:\Git\Golden Goose\Decrypt AES-128-ECB\TextToColumn.xlsx"
sheetName = 'Decrypt'

def removeChar(list):
    row = []
    for text in list:
        row.append(text.replace("@", ""))
    return row
    
#usando il delimitatore "!", i dati vengono spostati in colonna 
def textToColumn(inputFilePath,workbook): 
    rowIndex = 0
    #viene aggiunto uno sheet al woorkbook creato nel metodo main 
    worksheet = workbook.add_worksheet(sheetName)
    #il file sorgente originale viene aperto in formatok utf-16 per la gestione dei caratteri speciali
    with open(inputFilePath, newline='',encoding='utf-16') as input_file:
        #excel prevede un delimiter di un solo carattere. Prendo ogni riga, la gestisco come una stringa 
        #e uso il metodo split per estrarmi ij valori delle celle
        reader = csv.reader(input_file)
        for row in reader:
            splittedRow = row[0].split('@!@')
            #il valore processato viene inseito nella corrispettiva riga e colonna
            for i in range(len(splittedRow)):
                worksheet.write(rowIndex, i, splittedRow[i])
            rowIndex += 1

#rimuove tutti i caratteri speciai ascii e non quali \n, \t, \x06, \x10 ecc
def remove_control_chars(s):
    return ''.join(ch for ch in s if unicodedata.category(ch)[0] != 'C')

def decrypt(encyptedString):
    # Converto la chiave esadecimale in una sequenza di byte
    byteKey = bytes.fromhex(KEY)
    # Decodifico la stringa Base64 in una sequenza di byte
    encrypted_data = base64.b64decode(encyptedString)
    # Inizializzo il cifrario AES in modalità ECB con PKCS5Padding come padding
    cipher = AES.new(byteKey, AES.MODE_ECB)
    # Decripto i dati criptati
    decrypted_data = cipher.decrypt(encrypted_data)
    # Rimuovo i caratteri di controllo
    plaintext = remove_control_chars(decrypted_data.decode('utf-8'))
    # Restituisco il risultato
    return plaintext

#transcodifica data da formato unix timestamp a data
def fromUnixToDate(unix_timestamp): #esempio chiamata: print(fromUnixToDate(int(decrypt(cell.value))))
    data = datetime.datetime.fromtimestamp(unix_timestamp)
    data_str = data.strftime("%Y/%m/%d %H:%M:%S")
    return data_str

#restituisce il numero corrispondente alla letter. A = 1, B=2 ecc
def numFromLetter(letter):
    num = ord(letter) - ord('A') + 1
    return num 

#viene passato il file al path outputFile
def decryptFile(outputFile):
    #openpyxl serve pera prire il woorkbook
    workbook = openpyxl.load_workbook(outputFile)
    worksheet = workbook.get_sheet_by_name(sheetName)

    #la prima riga viene saltata perché coincide con il tracciato della tabella. Gestione con variabile booleana 
    for columns in columdToDecrypt:
        skip_first = True
        rowNumber = 2
        for cell in worksheet[columns]:
            if skip_first:
                skip_first = False
                continue  # salta la prima cella della prima colonna

            #attualmente convertito in formato data
            cellValue = worksheet.cell(row = rowNumber, column = numFromLetter(columns))
            #la colonna N è della data, quindi occorre applicare anche la funzione da timestamp unix a date
            #decripto ogni valore delle celle presenti in columdToDecrypt
            print(decrypt(cell.value))
            cellValue.value = decrypt(cell.value)
            rowNumber +=1
            workbook.save(filename=outputFile)

def main():
    #elimino il precedene file 
    if os.path.exists(outputFile):
        os.remove(outputFile)
    #creo il file excel e sposto il testo il varie colonne gestendo il delimiter @!@
    workbook = xlsxwriter.Workbook(outputFile)
    textToColumn(inputFile,workbook)
    workbook.close()
    #eseguo il processo di decrypt 
    decryptFile(outputFile)



if __name__ == '__main__':
    main()